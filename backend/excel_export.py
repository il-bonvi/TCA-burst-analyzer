from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.utils import fmt_time, safe_avg

# ──────────────────────────────────────────────────────────────────────────────
# COLOR PALETTE  ←  edit these to change all colors in the export
# Format: "FFRRGGBB"  (FF = fully opaque, then hex RGB)
# ──────────────────────────────────────────────────────────────────────────────

# Background / chrome
BG_DARK   = "FF0D0F14"   # very dark title bars
BG_MID    = "FF1A1D24"   # sheet title bars
BG_HEAD   = "FF252932"   # column header rows
BG_WHITE  = "FFFFFFFF"   # pure white cell bg (used for "All Bursts" alt rows)

# Metric accent colors  (match the HTML app)
C_WATT    = "FFC084FC"   # purple — power / watt values
C_HR      = "FFEF4444"   # red    — heart rate
C_CAD     = "FF4DA6FF"   # blue   — cadence
C_TIME    = "00B050"   # green  — time / duration
C_DELTA   = "FF3CCF7E"   # green  — delta above threshold
C_MAX     = "FF9555CC"   # deep purple — max power
C_MIN     = "FFD9A8FF"   # light lavender — min power
C_TOT     = "FFE91E63"   # pink-magenta — tot >= metrics

# Text colors
T_WHITE   = "FFFFFFFF"
T_DARK    = "FF0D0F14"
T_MUTED   = "FF6B7280"

# Border
BORDER_C  = "FFCCCCCC"
BORDER_H  = "FF444444"   # header borders (darker)


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _hex_to_argb(hex_color: str) -> str:
    return "FF" + hex_color.lstrip("#").upper()


def _get_duration_counts(
    duration_counts: dict | None, bursts: list
) -> dict[int, int]:
    """Ensure duration_counts is populated, reconstructing from bursts if needed."""
    if duration_counts:
        return duration_counts
    
    result: dict[int, int] = {}
    for b in bursts:
        dur = round(b["duration"])
        result[dur] = result.get(dur, 0) + 1
    return result


def _lighten(hex_color: str, factor: float) -> str:
    """Blend hex_color toward white by factor (0=unchanged, 1=white)."""
    h = hex_color.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r = round(r + (255 - r) * factor)
    g = round(g + (255 - g) * factor)
    b = round(b + (255 - b) * factor)
    return f"FF{r:02X}{g:02X}{b:02X}"


def _fill(argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=argb)


def _border(color: str = BORDER_C) -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _center(wrap: bool = False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(indent: int = 1) -> Alignment:
    return Alignment(horizontal="left", vertical="center", indent=indent)


def _font(
    color: str = T_DARK,
    bold: bool = False,
    size: int = 10,
    name: str = "Arial",
) -> Font:
    return Font(bold=bold, size=size, color=color, name=name)





def _write_header_row(ws, row: int, columns: list[tuple[str, int]]) -> None:
    """Write a styled header row and set column widths."""
    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font  = _font(color=T_WHITE, bold=True, size=10)
        cell.fill  = _fill(BG_HEAD)
        cell.alignment = _center()
        cell.border = _border(BORDER_H)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 20


def _write_title(ws, text: str, span: str, fill_argb: str, row: int = 1, title_color: str = T_WHITE) -> None:
    ws.merge_cells(span)
    cell = ws[span.split(":")[0]]
    cell.value = text
    cell.font  = _font(color=title_color, bold=True, size=12)
    cell.fill  = _fill(fill_argb)
    cell.alignment = _left(indent=1)
    ws.row_dimensions[row].height = 26


# ──────────────────────────────────────────────────────────────────────────────
# Main builder
# ──────────────────────────────────────────────────────────────────────────────

def build_excel(
    all_results: list[dict[str, Any]],
    min_dur: int = 4,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # ── SHEET 1: SUMMARY ──────────────────────────────────
    ws_sum = wb.create_sheet("Summary")

    ws_sum.merge_cells("A1:F1")
    c = ws_sum["A1"]
    c.value = "⚡  Burst Analysis — Summary"
    c.font  = _font(color=T_WHITE, bold=True, size=14)
    c.fill  = _fill(BG_DARK)
    c.alignment = _left()
    ws_sum.row_dimensions[1].height = 30

    SUM_COLS = [
        ("Threshold",    16),
        ("#",     11),
        ("t",   14),
        ("Pwr",    13),
        ("HR",       12),
        ("Cad",  14),
    ]
    _write_header_row(ws_sum, 2, SUM_COLS)
    
    # Fix Threshold header to be white text
    ws_sum["A2"].font = _font(color=T_WHITE, bold=True, size=10)

    for res in all_results:
        threshold = res["threshold"]
        color     = res.get("color", "#888888")
        bursts    = res.get("bursts", [])
        n = len(bursts)
        if n == 0:
            continue

        total_dur  = sum(b["duration"] for b in bursts)
        avg_power  = safe_avg([b["avg_power"] for b in bursts])
        avg_hr_v   = [b["avg_hr"] for b in bursts if b.get("avg_hr")]
        avg_cad_v  = [b["avg_cadence"] for b in bursts if b.get("avg_cadence")]
        avg_hr     = safe_avg(avg_hr_v)  if avg_hr_v  else None
        avg_cad    = safe_avg(avg_cad_v) if avg_cad_v else None

        bg_thr = _hex_to_argb(color)
        bg_row = _lighten(color, 0.82)

        row_num = ws_sum.max_row + 1
        values  = [
            f"≥ {round(threshold)} W",
            n,
            fmt_time(total_dur),
            f"{avg_power:.0f} W",
            f"{avg_hr:.0f} bpm"  if avg_hr  else "—",
            f"{avg_cad:.0f} rpm" if avg_cad else "—",
        ]
        text_colors = [T_DARK, T_DARK, C_TIME, C_WATT, C_HR, C_CAD]
        bolds       = [True,   True,   True,   True,   True, True]

        for col_idx, (val, tc, bold) in enumerate(zip(values, text_colors, bolds), start=1):
            cell = ws_sum.cell(row=row_num, column=col_idx, value=val)
            bg = bg_thr if col_idx == 1 else bg_row
            fg = T_DARK if col_idx == 1 else tc
            cell.fill      = _fill(bg)
            cell.font      = _font(color=fg, bold=bold, size=10)
            cell.alignment = _center()
            cell.border    = _border()
        ws_sum.row_dimensions[row_num].height = 19

    # ── SEPARATOR ROW ──────────────────────────────────────
    sep_row = ws_sum.max_row + 2   # 1 empty row gap
    ws_sum.row_dimensions[sep_row - 1].height = 10  # spacer row

    # ── TRANSPOSED ALL-THRESHOLDS GRID ────────────────────
    # Process results in DESCENDING threshold order
    ordered_results = sorted(all_results, key=lambda r: r["threshold"], reverse=True)

    # Collect all unique durations across all thresholds (>= min_dur), sorted ascending
    all_durations: list[int] = []
    for res in ordered_results:
        duration_counts = _get_duration_counts(res.get("duration_counts"), res.get("bursts", []))
        for d in duration_counts.keys():
            dur_int = int(d)
            if dur_int >= min_dur and dur_int not in all_durations:
                all_durations.append(dur_int)
    all_durations.sort()

    # Metric rows definition: (label, metric_key, color, format_fn)
    METRIC_ROWS = [
        ("Count",         "count",     C_TIME, lambda v, _: str(int(v)) if v else "—"),
        ("Total Time",    "tot_time",  C_TIME, lambda v, _: fmt_time(v) if v else "—"),
        ("Avg Power (W)", "avg_pow",   C_WATT, lambda v, _: f"{v:.0f}" if v else "—"),
        ("Avg HR (bpm)",  "avg_hr",    C_HR,   lambda v, _: f"{v:.0f}" if v else "—"),
        ("Avg Cad (rpm)", "avg_cad",   C_CAD,  lambda v, _: f"{v:.0f}" if v else "—"),
        ("Tot ≥ Count",   "cum_cnt",   C_TOT,  lambda v, _: str(int(v)) if v else "—"),
        ("Tot ≥ Time",    "cum_time",  C_TOT,  lambda v, _: fmt_time(v) if v else "—"),
    ]

    current_row = sep_row

    for res_idx, res in enumerate(ordered_results):
        threshold = res["threshold"]
        color     = res.get("color", "#888888")
        bursts    = res.get("bursts", [])
        duration_counts = _get_duration_counts(res.get("duration_counts"), bursts)

        argb    = _hex_to_argb(color)
        light   = _lighten(color, 0.70)
        vlight  = _lighten(color, 0.87)

        # Compute stats per duration for this threshold
        dur_stats: dict[int, dict] = {}
        for dur in all_durations:
            count = duration_counts.get(dur, duration_counts.get(str(dur), 0))
            if count == 0:
                # duration not present for this threshold
                dur_stats[dur] = None
                continue

            bd = [b for b in bursts if round(b["duration"]) == dur]
            total_d = sum(b["duration"] for b in bd)
            avg_pw  = safe_avg([b["avg_power"]   for b in bd])
            hr_v    = [b["avg_hr"]      for b in bd if b.get("avg_hr")]
            cad_v   = [b["avg_cadence"] for b in bd if b.get("avg_cadence")]
            avg_hr  = safe_avg(hr_v)  if hr_v  else None
            avg_cad = safe_avg(cad_v) if cad_v else None

            cum_b   = [b for b in bursts if round(b["duration"]) >= dur]
            cum_cnt = len(cum_b)
            cum_dur = sum(b["duration"] for b in cum_b)

            dur_stats[dur] = {
                "count":   count,
                "tot_time": total_d,
                "avg_pow":  avg_pw,
                "avg_hr":   avg_hr,
                "avg_cad":  avg_cad,
                "cum_cnt":  cum_cnt,
                "cum_time": cum_dur,
            }

        # ── Threshold label row (header with durations as columns) ──
        header_row = current_row

        # Col A: threshold label with soglia value (now in B)
        cell_b = ws_sum.cell(row=header_row, column=1, value=f"≥{round(threshold)} W")
        cell_b.font      = _font(color=T_DARK, bold=True, size=10)
        cell_b.fill      = _fill(argb)
        cell_b.alignment = _center()
        cell_b.border    = _border(BORDER_H)
        ws_sum.row_dimensions[header_row].height = 22

        # Cols B+: one column per duration
        for d_idx, dur in enumerate(all_durations):
            col_num = d_idx + 2  # starts at col B
            cell = ws_sum.cell(row=header_row, column=col_num, value=f"{dur}s")
            cell.font      = _font(color=T_WHITE, bold=True, size=10)
            cell.fill      = _fill(BG_HEAD)
            cell.alignment = _center()
            cell.border    = _border(BORDER_H)

        # ── Metric rows ──
        for m_idx, (m_label, m_key, m_color, m_fmt) in enumerate(METRIC_ROWS):
            data_row = header_row + 1 + m_idx
            row_bg   = light if m_key in ("count", "tot_time", "cum_cnt", "cum_time") else vlight

            # Col A: metric label
            cell_lbl = ws_sum.cell(row=data_row, column=1, value=m_label)
            cell_lbl.font      = _font(color=T_WHITE, bold=True, size=9)
            cell_lbl.fill      = _fill(BG_HEAD)
            cell_lbl.alignment = _left(indent=1)
            cell_lbl.border    = _border(BORDER_H)

            # Cols B+: values per duration
            for d_idx, dur in enumerate(all_durations):
                col_num  = d_idx + 2
                stats    = dur_stats.get(dur)
                val_raw  = stats[m_key] if stats else None
                val_str  = m_fmt(val_raw, dur) if val_raw is not None else "—"
                fg       = m_color if (m_color and val_raw is not None) else (T_MUTED if val_raw is None else T_DARK)
                cell = ws_sum.cell(row=data_row, column=col_num, value=val_str)
                cell.font      = _font(color=fg, bold=(val_raw is not None), size=10)
                cell.fill      = _fill(row_bg if val_raw is not None else vlight)
                cell.alignment = _center()
                cell.border    = _border()
            ws_sum.row_dimensions[data_row].height = 17

        current_row = header_row + len(METRIC_ROWS) + 1 + 1  # +1 for header row, +1 for spacer

        # Spacer row between thresholds
        if res_idx < len(ordered_results) - 1:
            ws_sum.row_dimensions[current_row - 1].height = 8

    # ── Set column widths for the transposed grid ──────────
    ws_sum.column_dimensions["A"].width = 14  # threshold label
    # Set summary columns B onward to 50px (Excel width ~7.2)
    for col_idx in range(2, ws_sum.max_column + 1):
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = 7.2

    ws_sum.freeze_panes = "A1"

    # ── SHEETS 2…N: PER-THRESHOLD DURATION GRID ───────────
    for res in all_results:
        threshold      = res["threshold"]
        color          = res.get("color", "#888888")
        bursts         = res.get("bursts", [])
        duration_counts = _get_duration_counts(res.get("duration_counts"), bursts)

        sheet_name = f"≥{round(threshold)}W"[:31]
        ws = wb.create_sheet(sheet_name)

        argb   = _hex_to_argb(color)
        light  = _lighten(color, 0.70)
        vlight = _lighten(color, 0.87)

        avg_pow_all = safe_avg([b["avg_power"] for b in bursts]) if bursts else 0
        avg_hr_all  = safe_avg([b["avg_hr"]  for b in bursts if b.get("avg_hr")])  or None
        avg_cad_all = safe_avg([b["avg_cadence"] for b in bursts if b.get("avg_cadence")]) or None
        total_time  = sum(b["duration"] for b in bursts)

        title_txt = (
            f"≥{round(threshold)} W  ·  {len(bursts)} bursts  ·  "
            f"Tot: {fmt_time(total_time)}  ·  "
            f"Avg Power: {avg_pow_all:.0f}W  ·  "
            f"Avg HR: {avg_hr_all:.0f} bpm" if avg_hr_all else
            f"≥{round(threshold)} W  ·  {len(bursts)} bursts  ·  "
            f"Tot: {fmt_time(total_time)}  ·  "
            f"Avg Power: {avg_pow_all:.0f}W"
        )
        _write_title(ws, title_txt, "A1:H1", argb)

        GRID_COLS = [
            ("Duration (s)",   13),
            ("Count",          10),
            ("Total Time",     13),
            ("Avg Power (W)",  14),
            ("Avg HR (bpm)",   14),
            ("Avg Cad (rpm)",  14),
            ("Tot ≥ Count",    13),
            ("Tot ≥ Time",     13),
        ]
        _write_header_row(ws, 2, GRID_COLS)

        sorted_durs = sorted(
            int(d) for d in duration_counts.keys() if int(d) >= min_dur
        )

        for dur in sorted_durs:
            count = duration_counts.get(dur, duration_counts.get(str(dur), 0))
            bd    = [b for b in bursts if round(b["duration"]) == dur]

            total_d  = sum(b["duration"] for b in bd)
            avg_pw   = safe_avg([b["avg_power"] for b in bd])
            hr_v     = [b["avg_hr"]       for b in bd if b.get("avg_hr")]
            cad_v    = [b["avg_cadence"]  for b in bd if b.get("avg_cadence")]
            avg_hr   = safe_avg(hr_v)  if hr_v  else None
            avg_cad  = safe_avg(cad_v) if cad_v else None

            cum_b   = [b for b in bursts if round(b["duration"]) >= dur]
            cum_cnt = len(cum_b)
            cum_dur = sum(b["duration"] for b in cum_b)

            row_vals   = [f"{dur}s", count, fmt_time(total_d),
                          f"{avg_pw:.0f}", f"{avg_hr:.0f}" if avg_hr else "—",
                          f"{avg_cad:.0f}" if avg_cad else "—",
                          cum_cnt, fmt_time(cum_dur)]

            bg_map   = {2: light, 3: light, 7: light, 8: light}
            tc_map   = {
                1: T_DARK,
                2: C_TIME,
                3: C_TIME,
                4: C_WATT,
                5: C_HR,
                6: C_CAD,
                7: C_TOT,
                8: C_TOT,
            }
            bold_map = {1, 2, 3, 4, 5, 6, 7, 8}

            row_num = ws.max_row + 1
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.fill      = _fill(bg_map.get(col_idx, vlight))
                cell.font      = _font(
                    color=tc_map.get(col_idx, T_DARK),
                    bold=(col_idx in bold_map),
                    size=10,
                )
                cell.alignment = _center()
                cell.border    = _border()
            ws.row_dimensions[row_num].height = 18

        ws.freeze_panes = "A3"

    # ── LAST SHEET: ALL BURSTS DETAIL ─────────────────────
    ws_all = wb.create_sheet("All Bursts")

    _write_title(ws_all, "📋  Tutti i Burst — Dettaglio Completo", "A1:J1", BG_MID, title_color=T_WHITE)

    ALL_COLS = [
        ("Threshold",     14),
        ("#",              6),
        ("Start",         10),
        ("Duration (s)",  13),
        ("Avg W",         11),
        ("Max W",         11),
        ("Min W",         11),
        ("Delta W",       11),
        ("HR avg",        11),
        ("Cad avg",       11),
    ]
    _write_header_row(ws_all, 2, ALL_COLS)

    for res in all_results:
        threshold = res["threshold"]
        color     = res.get("color", "#888888")
        vlight    = _lighten(color, 0.88)
        argb_thr  = _hex_to_argb(color)

        for b in res.get("bursts", []):
            hour = b.get("hour", "")
            if not hour or hour == "undefined":
                hour = fmt_time(b.get("start_time", 0))

            row_vals = [
                f"≥{round(threshold)}W",
                b["rank"],
                hour,
                b["duration"],
                round(b["avg_power"]),
                b["max_power"],
                b["min_power"],
                int(round(b.get("delta_above", 0))),
                b["avg_hr"]       if b.get("avg_hr")       else "—",
                b["avg_cadence"]  if b.get("avg_cadence")  else "—",
            ]
            tc_all = {
                1: T_DARK,
                2: T_MUTED,
                3: T_DARK,
                4: C_TIME,
                5: C_WATT,
                6: C_MAX,
                7: C_MIN,
                8: C_DELTA,
                9: C_HR,
                10: C_CAD,
            }
            bold_all = {1, 2, 3, 4, 5, 6, 7, 8, 9, 10}

            row_num = ws_all.max_row + 1
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws_all.cell(row=row_num, column=col_idx, value=val)
                cell.fill      = _fill(vlight)
                cell.font      = _font(
                    color=tc_all.get(col_idx, T_DARK),
                    bold=(col_idx in bold_all),
                    size=10,
                )
                cell.alignment = _center()
                cell.border    = _border()
            ws_all.row_dimensions[row_num].height = 17

    ws_all.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()